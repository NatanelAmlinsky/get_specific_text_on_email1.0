import json
import re
import openpyxl
import requests
import win32com.client
import datetime

class OutlookAccount:
    def __init__(self, account_name):
        self.message = None
        self.outlook = win32com.client.Dispatch("Outlook.Application")
        self.namespace = self.outlook.GetNamespace("MAPI")
        self.account_name = account_name
        self.account = None

    def login(self):
        for a in self.namespace.Accounts:
            if a.DisplayName == self.account_name:
                self.account = a
                break

        if not self.account:
            print(f"Could not find account '{self.account_name}'")
            return False
        self.inbox_folder = self.account.DeliveryStore.GetDefaultFolder(6)
        return True

    def get_table_content(self, email_body):
        # Split the email body into lines
        self.email_body = email_body
        lines = email_body.split("\n")

        # Identify the table rows based on patterns or characteristics
        table_rows = []
        for line in lines:
            table_rows.append(line.strip())

        # Combine the table rows into a single string
        table_content = "\n".join(table_rows)
        return table_content

    def get_email_content(self, message):
        json_file = open('configuration.json', 'r', encoding='utf-8')  # open the JSON file
        data = json.load(json_file)  # loading the JSON content to the variable "data"
        self.message = message
        email_body = self.message.body
        subject = self.message.Subject
        print(subject)
        # Search for the table pattern in the email body
        table_pattern = r'\w+:\s*\w+'  # Modify the pattern as per your table format
        table_match = re.search(table_pattern, email_body)

        if table_match:
            # If table pattern is found, extract the table content
            table_content = self.get_table_content(email_body)
            email_body = table_content
            # Process the table content as desired

        # Get the timestamp of the message
        timestamp_str = self.message.CreationTime.strftime('%d/%m/%Y %I:%M %p')
        # Parse the timestamp string into a Python datetime object
        timestamp = datetime.datetime.strptime(timestamp_str, '%d/%m/%Y %I:%M %p')

        # Format the date and time components separately
        date_str = timestamp.strftime('%d/%m/%Y')
        time_str = timestamp.strftime('%H:%M')
        parts = []

        order_info = {"Organization Name": "", "Order Number": "", "First Name": "",
                      "Last Name": "", "Full Name": "", "City": "", "Country": "",
                      "Land": "","House Number": "","Apartment Number": "", "Address": "", "Email": "",
                      "Phone Number": "", "Books": "", "Book Language": "", "Contact Me": "",
                      "Up 18": "", "Message": "",  "IP Address": "", "Birthday Year": "",
                      "Background": "", "Zip Code": ""}

        for line in email_body.splitlines():
            parts.extend(line.split("\n"))

        # Identify what organization name sent the order using JSON file up to location in the Organization Name list
        for index, sender_email in enumerate(data["senders email"]):
            if message.SenderEmailAddress == sender_email:
                order_info["Organization Name"] = data["Organization Name"][index]
        count = 0
        for element in parts:

            for order_number_txt in data["Order Number"]:
                if order_number_txt in data["Order Number"]:
                    if order_number_txt in element:
                        order_info["Order Number"] = element.split(":")[1]

            for full_name_txt in data["Full Name"]:
                if full_name_txt in data["Full Name"]:
                    if full_name_txt in element:
                        order_info["Full Name"] = element.split(":")[1]
                        order_info["First Name"] = "-"
                        order_info["Last Name"] = "-"

            for first_name in data["First Name"]:
                if first_name in data["First Name"]:
                    if first_name in element:
                        order_info["First Name"] = element.split(":")[1]

            for last_name in data["Last Name"]:
                if last_name in data["Last Name"]:
                    if last_name in element:
                        order_info["Last Name"] = element.split(":")[1]
                        order_info["Full Name"] = order_info["First Name"] + " " + order_info["Last Name"]

            for address_txt in data["Address"]:
                if address_txt in data["Address"]:
                    if address_txt in element:
                        order_info["Address"] = element.split(":")[1]

            for city_txt in data["City"]:
                if city_txt in data["City"]:
                    if city_txt in element:
                        order_info["City"] = element.split(":")[1]
            for country_txt in data["Country"]:
                if country_txt in data["Country"]:
                    if country_txt in element:
                        order_info["Country"] = element.split(":")[1]

            for land_txt in data["Land"]:
                if land_txt in data["Land"]:
                    if land_txt in element:
                        order_info["Land"] = element.split(":")[1]

            for up_18_txt in data["Up 18"]:
                if up_18_txt in data["Up 18"]:
                    if up_18_txt in element:
                        order_info["Up 18"] = element.split(":")[1]

            for email_txt in data["Email"]:
                if email_txt in data["Email"]:
                    if email_txt in element:
                        order_info["Email"] = element.split(":")[1]
                        email = order_info["Email"]
                        fixed_email = email.split("<")[0]
                        order_info["Email"] = fixed_email

            for phone_number_txt in data["Phone Number"]:
                if phone_number_txt in data["Phone Number"]:
                    if phone_number_txt in element:
                        order_info["Phone Number"] = element.split(":")[1]

            for contact_me_txt in data["Contact Me"]:
                if contact_me_txt in data["Contact Me"]:
                    if contact_me_txt in element:
                        order_info["Contact Me"] = element.split(":")[1]

            for chosen_books_txt in data["Chosen Books"]:
                if chosen_books_txt in data["Chosen Books"]:
                    if chosen_books_txt in element:
                        order_info["Books"] = element.split(":")[1]

            # If statement for unusual body text for unusual emails
            if order_info["Organization Name"] == "JewishTestimonies":
                for chosen_books_txt2 in data["Unusual Chosen Books"]:
                    if chosen_books_txt2 in data["Unusual Chosen Books"]:
                        if chosen_books_txt2 in element:
                            order_info["Books"] = order_info["Books"] + ", " + element

            if order_info["Organization Name"] == "Yeshua4U":
                for y4u_books in data["Yeshua4U Unusual Books"]:
                    if y4u_books in data["Yeshua4U Unusual Books"]:
                        if y4u_books in subject:
                            order_info["Books"] = subject.split("- ")[1]

            if order_info["Organization Name"] == "Medabrim":
                for chosen_books_txt3 in data["Unusual Chosen Books"]:
                    if chosen_books_txt3 in data["Unusual Chosen Books"]:
                        if chosen_books_txt3 in element:
                            if "Age Confirmation" not in element:
                                medabrim_book = element.split(":")[0]
                                order_info["Books"] = order_info["Books"] + ", " + medabrim_book

            for apartment_number_txt in data["Apartment Number"]:
                if apartment_number_txt in data["Apartment Number"]:
                    if apartment_number_txt in element:
                        order_info["Apartment Number"] = element.split(":")[1]

            for house_number_txt in data["House Number"]:
                if house_number_txt in data["House Number"]:
                    if house_number_txt in element:
                        order_info["House Number"] = element.split(":")[1]

            for ip_address_txt in data["IP Address"]:
                if ip_address_txt in data["IP Address"]:
                    if ip_address_txt in element:
                        order_info["IP Address"] = element.split(":")[1]
                        ip_address = order_info["IP Address"]
            for book_language_txt in data["Book Language"]:
                if book_language_txt in data["Book Language"]:
                    if book_language_txt in element:
                        order_info["Book Language"] = element.split(":")[1]
                        if order_info["Organization Name"] == "OneForIsrael":
                            order_info["Books"] = "ברית חדשה"

            for background_txt in data["Background"]:
                if background_txt in data["Background"]:
                    if background_txt in element:
                        order_info["Background"] = element.split(":")[1]

            for birthday_year_txt in data["Birthday Year"]:
                if birthday_year_txt in data["Birthday Year"]:
                    if birthday_year_txt in element:
                        order_info["Birthday Year"] = element.split(":")[1]

            for message_txt in data["Message"]:
                if message_txt in data["Message"]:
                    if message_txt in element:
                        try:
                            # Remove non-ASCII characters from the string
                            # In this if statement I discovered that to grab the message content
                            # I had to understand that I am dealing with 2 lines.
                            # Every element focusing in one specific line.
                            # But I fixed it by go to the next line after "תוכן ההודעה:" to extract the content
                            cleaned_message = re.sub(r'[^\w\s]', '', element.split(":")[1])
                            order_info["Message"] = cleaned_message

                            if order_info["Message"] == "":
                                index = parts.index(element)
                                if index < len(parts) - 1:
                                    # Get the next line and remove leading/trailing whitespace
                                    message_content = parts[index + 1].strip()
                                else:
                                    message_content = ""
                                order_info["Message"] = message_content
                        except:
                            index = parts.index(element)
                            if index < len(parts) - 1:
                                # Get the next line and remove leading/trailing whitespace
                                message_content = parts[index + 1].strip()
                            else:
                                message_content = ""
                            order_info["Message"] = message_content

            for zip_code_txt in data["Zip Code"]:
                if zip_code_txt in data["Zip Code"]:
                    if zip_code_txt in element:
                        order_info["Zip Code"] = element.split(":")[1]

        # Create a new workbook and select the active worksheet
        if order_info["First Name"] == "" or order_info["Last Name"] == "":
            full_name = order_info["First Name"] + " " + order_info["Last Name"]
            order_info["Full Name"] = full_name

        wb = openpyxl.load_workbook("C:\\Users\\natan\\Desktop\\EmailAutomation\\orders_info.xlsx")
        ws = wb.active

        # Write the headers to the first row of the worksheet if the worksheet is empty
        if not any(ws.iter_rows()):
            headers = ["Organization", "Time", "Date", "Order Number", "Full Name","House Number", "Apartment Number",
                       "Address", "Land", "Country", "City", "Zip Code", "Email", "Phone Number", "Books",
                       "Book Language", "Background", "Birthday Year", "Contact Me", "Message", "IP Address",
                       "Email Link"]
            ws.append(headers)
            # Get the email link
        entry_id = message.EntryID
        email_link = f'outlook:{entry_id}'

        # Find the first empty row
        current_row = 2
        while ws.cell(row=current_row, column=1).value is not None:
            current_row += 1

        # Write the order info to the empty row
        row = [order_info["Organization Name"],
               time_str,
               date_str,
               order_info["Order Number"],
               order_info["Full Name"],
               order_info["House Number"],
               order_info["Apartment Number"],
               order_info["Address"],
               order_info["Land"],
               order_info["Country"],
               order_info["City"],
               order_info["Zip Code"],
               order_info["Email"],
               order_info["Phone Number"],
               order_info["Books"],
               order_info["Book Language"],
               order_info["Background"],
               order_info["Birthday Year"],
               order_info["Contact Me"],
               order_info["Message"],
               order_info["IP Address"],
               email_link]

        for i, value in enumerate(row):
            ws.cell(row=current_row, column=i + 1, value=value)

        # Save the workbook to a file
        wb.save("C:\\Users\\natan\\Desktop\\EmailAutomation\\orders_info.xlsx")
        return print(order_info)


